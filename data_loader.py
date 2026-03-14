"""
data_loader.py
Reads the user's behavioral Excel file and structures it for the agent.
"""

import openpyxl
from datetime import time


def load_user_data(filepath: str) -> dict:
    """
    Load all sheets from the Excel file and return as structured dicts.
    Returns:
        {
            "signals":   list of signal dicts  (Signal Theory sheet),
            "loops":     list of loop dicts    (Loops sheet),
            "decisions": list of decision dicts (Decision Log sheet),
            "patterns":  list of pattern dicts  (Pattern Engine sheet),
        }
    """
    wb = openpyxl.load_workbook(filepath)

    signals   = _load_signals(wb["Signal theory"])
    loops     = _load_simple_sheet(wb["Loops"])
    decisions = _load_simple_sheet(wb["Decision Log"])
    patterns  = _load_simple_sheet(wb["Pattern Engine"])

    return {
        "signals":   signals,
        "loops":     loops,
        "decisions": decisions,
        "patterns":  patterns,
    }


def _load_signals(ws) -> list:
    """
    Parse the Signal Theory sheet.
    Rows alternate between: day-label rows (e.g. 'Monday') and event rows.
    Returns a list of dicts with a 'Day' key added.
    """
    headers = None
    current_day = "Unknown"
    records = []

    for row in ws.iter_rows(values_only=True):
        # Skip completely empty rows
        if all(v is None for v in row):
            continue

        # Grab header row
        if row[0] == "Time" and headers is None:
            headers = [h for h in row if h is not None]
            continue

        # Day label row (e.g. 'Monday', 'Tuesday', …)
        if isinstance(row[0], str) and row[1] is None:
            current_day = row[0]
            continue

        # Data row — must have a time value in column 0
        if isinstance(row[0], time):
            record = {"Day": current_day}
            for i, header in enumerate(headers):
                value = row[i] if i < len(row) else None
                if isinstance(value, time):
                    value = value.strftime("%H:%M")
                record[header] = value
            records.append(record)

    return records


def _load_simple_sheet(ws) -> list:
    """
    Load a sheet that has a clean header row followed by data rows.
    Skips rows where the first cell is None.
    """
    headers = None
    records = []

    for row in ws.iter_rows(values_only=True):
        # Skip empty rows
        if all(v is None for v in row):
            continue

        if headers is None:
            # First non-empty row is the header
            headers = [str(h) if h is not None else f"Col_{i}"
                       for i, h in enumerate(row)]
            continue

        # Only include rows that have a value in the first column
        if row[0] is None:
            continue

        record = {}
        for i, header in enumerate(headers):
            value = row[i] if i < len(row) else None
            record[header] = value

        records.append(record)

    return records


def format_data_summary(user_data: dict) -> str:
    """
    Convert the structured user data into a readable text block
    that can be injected into tool responses.
    """
    lines = []

    # ── Signal Theory ──────────────────────────────────────────────────
    signals = user_data.get("signals", [])
    lines.append(f"=== WEEKLY ACTIVITY LOG ({len(signals)} events) ===")
    current_day = None
    for s in signals:
        if s.get("Day") != current_day:
            current_day = s.get("Day")
            lines.append(f"\n--- {current_day} ---")
        lines.append(
            f"  {s.get('Time', '??:??')} | {s.get('Event', '?')} "
            f"| Type: {s.get('Type', '?')} "
            f"| Emotion: {s.get('Emotion', '?')} "
            f"| Signal: {s.get('Signal', '?')}"
        )

    # ── Loops ──────────────────────────────────────────────────────────
    loops = user_data.get("loops", [])
    lines.append(f"\n\n=== OPEN/CLOSED LOOPS ({len(loops)} loops) ===")
    for l in loops:
        lines.append(
            f"  [{l.get('Loop ID')}] {l.get('Loop')} | "
            f"Type: {l.get('Type')} | Status: {l.get('Status')} | "
            f"Mentions: {l.get('Mentions')} | "
            f"Opened: {l.get('Opened')} | Closed: {l.get('Closed')} | "
            f"Notes: {l.get('Notes')}"
        )

    # ── Decision Log ───────────────────────────────────────────────────
    decisions = user_data.get("decisions", [])
    lines.append(f"\n\n=== DECISION LOG ({len(decisions)} decisions) ===")
    for d in decisions:
        lines.append(
            f"  [{d.get('Time')}] {d.get('Decision')} | "
            f"Context: {d.get('Context')} | "
            f"Reasoning: {d.get('Reasoning')} | "
            f"Emotion: {d.get('Emotion')} | "
            f"Confidence: {d.get('Confidence')} | "
            f"Outcome: {d.get('Outcome')}"
        )

    # ── Pattern Engine ─────────────────────────────────────────────────
    patterns = user_data.get("patterns", [])
    lines.append(f"\n\n=== PATTERN ENGINE INSIGHTS ({len(patterns)} insights) ===")
    for p in patterns:
        lines.append(
            f"  [{p.get('Insight ID')} | {p.get('Pattern Type')} | "
            f"Confidence: {p.get('Confidence')}]\n"
            f"    Observation : {p.get('Observation')}\n"
            f"    Insight     : {p.get('Insight')}\n"
            f"    Reflection  : {p.get('Reflection Prompt')}"
        )

    return "\n".join(lines)
